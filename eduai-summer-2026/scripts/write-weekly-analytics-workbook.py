#!/usr/bin/env python3

import json
import shutil
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, NumData, NumVal, StrData, StrRef, StrVal
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


WEEK_ONE_START = date(2026, 5, 4)
DEFAULT_EXPECTED_HOURS = 20
DEFAULT_BASE_HOURS = 3

NAME_MAP = {
    "ariqmuldi": "Ariq Muldi",
    "Whiteknight07": "Stavan",
    "evanbones": "Evan Bowness",
    "yta3216": "Ye",
    "abdullahmoh21": "Abdullah Mohsin Naqvi",
    "GlowyBlack": "Al-Ameen",
    "Ayyhab": "Ahab Masud",
    "superbolt08": "Syed Saad Ali",
    "ebabar5": "Ehsan",
    "KitheK": "Kithe",
    "frostbitcactus": "Ammaar",
}


def usage():
    print(
        "Usage: write-weekly-analytics-workbook.py <weekly-hours-with-pr-analytics.json> [output.xlsx]",
        file=sys.stderr,
    )
    sys.exit(2)


def week_start(week_number):
    return WEEK_ONE_START + timedelta(days=(week_number - 1) * 7)


def week_label(week_number):
    start = week_start(week_number)
    return f"Week {week_number} ({start.strftime('%b')} {start.day})"


def week_dates(week_number):
    start = week_start(week_number)
    end = start + timedelta(days=6)
    if start.month == end.month:
        return f"{start.strftime('%B')} {start.day}-{end.day}, {start.year}"
    return f"{start.strftime('%B')} {start.day}-{end.strftime('%B')} {end.day}, {start.year}"


def sheet_name_for_week(week_number):
    return week_label(week_number).replace("Jun ", "Jun ").replace("May ", "May ")


def display_number(value):
    if value in ("", None):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    return int(number) if number.is_integer() else round(number, 2)


def number_cache(values):
    return NumData(
        formatCode="0.0",
        ptCount=len(values),
        pt=[NumVal(idx=index, v=float(value)) for index, value in enumerate(values)],
    )


def string_cache(values):
    return StrData(
        ptCount=len(values),
        pt=[StrVal(idx=index, v=str(value)) for index, value in enumerate(values)],
    )


def cached_number(value):
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.6f}".rstrip("0").rstrip(".")


def add_table(sheet, table_name):
    if sheet.max_row < 2:
        return
    ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def hyperlink_formula(url, label):
    if not url:
        return ""
    escaped_url = str(url).replace('"', '""')
    escaped_label = str(label).replace('"', '""')
    return f'=HYPERLINK("{escaped_url}","{escaped_label}")'


def roster_from_data(data):
    seen = []
    for row in data.get("detailRows", []):
        username = row.get("github_username")
        if username and username not in seen:
            seen.append(username)
    ordered = [username for username in NAME_MAP if username in seen]
    ordered.extend(username for username in NAME_MAP if username not in ordered)
    ordered.extend(username for username in seen if username not in ordered)
    return ordered


def aggregate_rows(data):
    hours = defaultdict(float)
    issue_sets = defaultdict(set)

    for row in data.get("detailRows", []):
        username = row.get("github_username")
        week = row.get("week")
        if not username or not week:
            continue
        key = (week, username)
        hours[key] += float(row.get("counted_implementation_hours", row.get("implementation_hours", 0)) or 0)
        if row.get("issue_number"):
            issue_sets[key].add(row["issue_number"])

    return {
        key: {
            "hours": round(value, 2),
            "issue_count": len(issue_sets[key]),
        }
        for key, value in hours.items()
    }


def style_workbook(workbook):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    manual_fill = PatternFill("solid", fgColor="FCE4D6")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if sheet.title == "Manual Review":
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = manual_fill

        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 11), 42)


def add_readme(workbook, data, weeks):
    sheet = workbook.active
    sheet.title = "README"
    sheet.append(["Topic", "Details"])
    sheet.append(["What this is", "Professor-facing weekly implementation-hours tracker."])
    sheet.append(["Repository", f"{data['context']['owner']}/{data['context']['repo']}"])
    sheet.append(["Reporting period", f"{data['context']['report_start_date']} to {data['context']['report_end_date']}"])
    sheet.append(["Issue scope", data["context"].get("project_scope", "Project 8 Done/closed issues only.")])
    sheet.append(["Weeks included", ", ".join(f"{week_label(week)}: {week_dates(week)}" for week in weeks)])
    sheet.append(["Expected hours", "Default is 20 hours per person per week; change Settings!B2 to update deltas."])
    sheet.append(["Base hours", "Default is 3 hours per person per week; change Settings!B3 or override person/week cells in Base Hours."])
    sheet.append(["Actual hours", "Issue-body implementation hours from completed Project 8 issues."])
    sheet.append(["Workbook shape", "Each week sheet has exactly one row per person. Issue-level detail sheets are intentionally omitted."])
    sheet.append(["PR analytics", "Not included."])


def add_settings(workbook, data):
    sheet = workbook.create_sheet("Settings")
    sheet.append(["Setting", "Value", "Notes"])
    sheet.append(["Expected hours per person per week", DEFAULT_EXPECTED_HOURS, "Edit this value to update all expected/delta formulas."])
    sheet.append(["Default base hours per person per week", DEFAULT_BASE_HOURS, "Meeting/admin/default weekly hours. Edit this or override specific cells in Base Hours."])
    sheet.append(["Reporting period", f"{data['context']['report_start_date']} to {data['context']['report_end_date']}", "Generated from Project 8 Done/closed issues."])
    sheet.append(["Project scope", data["context"].get("project_scope", ""), f"{data['context'].get('project_owner', '')} project {data['context'].get('project_number', '')}"])
    sheet.append(["Generated at", data["context"].get("generated_at", ""), "UTC timestamp from the report generator."])


def add_people(workbook, roster):
    sheet = workbook.create_sheet("People")
    sheet.append(["Actual Name", "GitHub Handle", "Needs Name Confirmation"])
    for username in roster:
        actual_name = NAME_MAP.get(username, "")
        sheet.append([actual_name, username, "YES" if not actual_name else ""])
    add_table(sheet, "PeopleTable")


def add_base_hours(workbook, weeks, roster):
    sheet = workbook.create_sheet("Base Hours")
    sheet.append(["Actual Name", "GitHub Handle", *[week_label(week) for week in weeks]])
    for username in roster:
        sheet.append([NAME_MAP.get(username, ""), username, *["=Settings!$B$3" for _ in weeks]])
    add_table(sheet, "BaseHours")


def add_week_sheet(workbook, week_number, roster, aggregation):
    label = week_label(week_number)
    sheet = workbook.create_sheet(sheet_name_for_week(week_number))
    sheet.append(
        [
            "Week",
            "Week Dates",
            "Actual Name",
            "GitHub Handle",
            "Expected Hours",
            "Implementation Hours",
            "Base Hours",
            "Total Actual Hours",
            "Implementation Issue Count",
            "Delta Total-Expected",
        ]
    )
    base_col = get_column_letter(week_number + 2)
    for roster_index, username in enumerate(roster, start=2):
        row_index = sheet.max_row + 1
        actual = aggregation.get((f"Week {week_number}", username), {"hours": 0, "issue_count": 0})
        sheet.append(
            [
                label,
                week_dates(week_number),
                NAME_MAP.get(username, ""),
                username,
                "=Settings!$B$2",
                actual["hours"],
                f"='Base Hours'!{base_col}{roster_index}",
                f"=F{row_index}+G{row_index}",
                actual["issue_count"],
                f"=H{row_index}-E{row_index}",
            ]
        )
    add_table(sheet, f"Week{week_number}Summary")


def add_weekly_summary(workbook, weeks, roster):
    sheet = workbook.create_sheet("Weekly Summary")
    sheet.append(
        [
            "Week",
            "Week Dates",
            "Actual Name",
            "GitHub Handle",
            "Expected Hours",
            "Implementation Hours",
            "Base Hours",
            "Total Actual Hours",
            "Implementation Issue Count",
            "Delta Total-Expected",
        ]
    )

    for week in weeks:
        week_sheet = sheet_name_for_week(week)
        for offset, username in enumerate(roster, start=2):
            row_index = sheet.max_row + 1
            sheet.append(
                [
                    week_label(week),
                    week_dates(week),
                    NAME_MAP.get(username, ""),
                    username,
                    f"='{week_sheet}'!E{offset}",
                    f"='{week_sheet}'!F{offset}",
                    f"='{week_sheet}'!G{offset}",
                    f"='{week_sheet}'!H{offset}",
                    f"='{week_sheet}'!I{offset}",
                    f"='{week_sheet}'!J{offset}",
                ]
            )
    add_table(sheet, "WeeklySummary")


def add_chart(workbook, weeks, roster, aggregation):
    sheet = workbook.create_sheet("Charts")
    sheet.append(["Expected vs Actual Hours Dashboard"])
    sheet.append(["Readable chart sheet: two large charts plus exact values in the tables below."])
    sheet.append(["Chart 1: team member on x-axis, total hours on y-axis."])
    sheet.cell(42, 1).value = "Chart 2: week on x-axis, total actual hours on y-axis."

    summary_header = 80
    sheet.cell(summary_header - 1, 1).value = "Summary used by Chart 1"
    sheet.cell(summary_header, 1).value = "Team Member"
    sheet.cell(summary_header, 2).value = "Expected Hours"
    sheet.cell(summary_header, 3).value = "Actual Hours"
    sheet.cell(summary_header, 4).value = "Variance"
    sheet.cell(summary_header, 5).value = "Completion %"

    team_names = []
    expected_totals = []
    actual_totals = []
    row_index = summary_header + 1
    for username in roster:
        implementation_total = sum(
            aggregation.get((f"Week {week}", username), {"hours": 0})["hours"]
            for week in weeks
        )
        expected_total = DEFAULT_EXPECTED_HOURS * len(weeks)
        actual_total = implementation_total + DEFAULT_BASE_HOURS * len(weeks)
        team_names.append(NAME_MAP.get(username, username))
        expected_totals.append(expected_total)
        actual_totals.append(round(actual_total, 2))

        sheet.cell(row_index, 1).value = NAME_MAP.get(username, username)
        sheet.cell(row_index, 2).value = f"=Settings!$B$2*{len(weeks)}"
        sheet.cell(row_index, 3).value = (
            f'=SUMIFS(WeeklySummary[Total Actual Hours],'
            f'WeeklySummary[GitHub Handle],"{username}")'
        )
        sheet.cell(row_index, 4).value = f"=C{row_index}-B{row_index}"
        sheet.cell(row_index, 5).value = f"=IFERROR(C{row_index}/B{row_index},0)"
        row_index += 1

    weekly_header = 80
    sheet.cell(weekly_header - 1, 7).value = "Weekly totals used by Chart 2"
    sheet.cell(weekly_header, 7).value = "Week"
    sheet.cell(weekly_header, 8).value = "Total Actual Hours"
    week_names = []
    week_totals = []
    week_total_row = weekly_header + 1
    for week in weeks:
        implementation_total = sum(
            aggregation.get((f"Week {week}", username), {"hours": 0})["hours"]
            for username in roster
        )
        week_total = implementation_total + DEFAULT_BASE_HOURS * len(roster)
        week_names.append(week_label(week))
        week_totals.append(round(week_total, 2))

        sheet.cell(week_total_row, 7).value = week_label(week)
        sheet.cell(week_total_row, 8).value = (
            f'=SUMIFS(WeeklySummary[Total Actual Hours],'
            f'WeeklySummary[Week],G{week_total_row})'
        )
        week_total_row += 1

    source_header = 103
    sheet.cell(source_header - 1, 1).value = "Original source data"
    sheet.cell(source_header, 1).value = "Actual Name"
    sheet.cell(source_header, 2).value = "Week"
    sheet.cell(source_header, 3).value = "Expected Hours"
    sheet.cell(source_header, 4).value = "Actual Hours"

    source_row = source_header + 1
    for username in roster:
        for week in weeks:
            sheet.cell(source_row, 1).value = NAME_MAP.get(username, username)
            sheet.cell(source_row, 2).value = week_label(week)
            sheet.cell(source_row, 3).value = "=Settings!$B$2"
            sheet.cell(source_row, 4).value = (
                f'=SUMIFS(WeeklySummary[Total Actual Hours],'
                f'WeeklySummary[GitHub Handle],"{username}",WeeklySummary[Week],B{source_row})'
            )
            source_row += 1

    chart_one = BarChart()
    chart_one.type = "col"
    chart_one.style = 2
    chart_one.title = "Total Expected vs Actual Hours by Team Member"
    chart_one.x_axis.title = "Team Member"
    chart_one.y_axis.title = "Hours"
    chart_one.legend.position = "b"
    chart_one.dataLabels = DataLabelList()
    chart_one.dataLabels.showVal = False
    chart_one.dispBlanksAs = "zero"
    chart_one.x_axis.axPos = "b"
    chart_one.y_axis.axPos = "l"
    chart_one.y_axis.numFmt = "0.0"
    chart_one.add_data(
        Reference(sheet, min_col=2, max_col=3, min_row=summary_header, max_row=row_index - 1),
        titles_from_data=True,
    )
    chart_one.set_categories(Reference(sheet, min_col=1, min_row=summary_header + 1, max_row=row_index - 1))
    for index, series in enumerate(chart_one.series):
        series.cat = AxDataSource(strRef=StrRef(f=f"'Charts'!$A${summary_header + 1}:$A${row_index - 1}"))
        series.cat.strRef.strCache = string_cache(team_names)
        series.val.numRef.numCache = number_cache(expected_totals if index == 0 else actual_totals)
        if series.tx and series.tx.strRef:
            series.tx.strRef.strCache = string_cache(["Expected Hours" if index == 0 else "Actual Hours"])
    chart_one.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=0, row=3),
        to=AnchorMarker(col=26, row=39),
    )
    sheet.add_chart(chart_one)

    chart_two = BarChart()
    chart_two.type = "col"
    chart_two.style = 2
    chart_two.title = "Total Actual Hours by Week"
    chart_two.x_axis.title = "Week"
    chart_two.y_axis.title = "Total Actual Hours"
    chart_two.legend = None
    chart_two.dataLabels = DataLabelList()
    chart_two.dataLabels.showVal = False
    chart_two.dispBlanksAs = "zero"
    chart_two.x_axis.axPos = "b"
    chart_two.y_axis.axPos = "l"
    chart_two.y_axis.numFmt = "0.0"
    chart_two.add_data(
        Reference(sheet, min_col=8, min_row=weekly_header, max_row=week_total_row - 1),
        titles_from_data=True,
    )
    chart_two.set_categories(Reference(sheet, min_col=7, min_row=weekly_header + 1, max_row=week_total_row - 1))
    for series in chart_two.series:
        series.cat = AxDataSource(strRef=StrRef(f=f"'Charts'!$G${weekly_header + 1}:$G${week_total_row - 1}"))
        series.cat.strRef.strCache = string_cache(week_names)
        series.val.numRef.numCache = number_cache(week_totals)
        if series.tx and series.tx.strRef:
            series.tx.strRef.strCache = string_cache(["Total Actual Hours"])
    chart_two.anchor = TwoCellAnchor(
        _from=AnchorMarker(col=0, row=42),
        to=AnchorMarker(col=26, row=73),
    )
    sheet.add_chart(chart_two)

    for display_name, ref in (
        ("ChartTeamTotals", f"A{summary_header}:E{row_index - 1}"),
        ("ChartWeeklyTotals", f"G{weekly_header}:H{week_total_row - 1}"),
        ("ChartSourceData", f"A{source_header}:D{source_row - 1}"),
    ):
        table = Table(displayName=display_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def add_manual_review(workbook, data):
    sheet = workbook.create_sheet("Manual Review")
    sheet.append(["Issue #", "Issue Link", "Issue Title", "Status", "Source Line", "Note"])
    owner = data["context"]["owner"]
    repo = data["context"]["repo"]
    for row in data.get("manualReviewRows", []):
        issue_number = row.get("issue_number", "")
        issue_url = f"https://github.com/{owner}/{repo}/issues/{issue_number}" if issue_number else ""
        sheet.append(
            [
                issue_number,
                hyperlink_formula(issue_url, f"#{issue_number}") if issue_number else "",
                row.get("issue_title", ""),
                row.get("status", ""),
                row.get("source_line", ""),
                row.get("note", ""),
            ]
        )
    add_table(sheet, "ManualReview")


def implementation_for(aggregation, week, username):
    return aggregation.get((f"Week {week}", username), {"hours": 0, "issue_count": 0})


def workbook_sheet_paths(xlsx_path):
    spreadsheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    relationships_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_relationships_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

    with ZipFile(xlsx_path) as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall(f"{{{package_relationships_ns}}}Relationship")
    }

    sheet_paths = {}
    for sheet in workbook_root.find(f"{{{spreadsheet_ns}}}sheets"):
        sheet_name = sheet.attrib["name"]
        relationship_id = sheet.attrib[f"{{{relationships_ns}}}id"]
        target = rel_targets[relationship_id]
        target = target.lstrip("/")
        sheet_paths[sheet_name] = target if target.startswith("xl/") else f"xl/{target}"
    return sheet_paths


def cached_formula_replacements(weeks, roster, aggregation):
    replacements = defaultdict(dict)

    for roster_index, username in enumerate(roster, start=2):
        for week in weeks:
            column = get_column_letter(week + 2)
            replacements["Base Hours"][f"{column}{roster_index}"] = DEFAULT_BASE_HOURS

    for week in weeks:
        sheet_name = sheet_name_for_week(week)
        for row_index, username in enumerate(roster, start=2):
            actual = implementation_for(aggregation, week, username)
            implementation_hours = actual["hours"]
            total_hours = implementation_hours + DEFAULT_BASE_HOURS
            replacements[sheet_name][f"E{row_index}"] = DEFAULT_EXPECTED_HOURS
            replacements[sheet_name][f"G{row_index}"] = DEFAULT_BASE_HOURS
            replacements[sheet_name][f"H{row_index}"] = total_hours
            replacements[sheet_name][f"J{row_index}"] = total_hours - DEFAULT_EXPECTED_HOURS

    weekly_summary_row = 2
    for week in weeks:
        for username in roster:
            actual = implementation_for(aggregation, week, username)
            implementation_hours = actual["hours"]
            total_hours = implementation_hours + DEFAULT_BASE_HOURS
            replacements["Weekly Summary"][f"E{weekly_summary_row}"] = DEFAULT_EXPECTED_HOURS
            replacements["Weekly Summary"][f"F{weekly_summary_row}"] = implementation_hours
            replacements["Weekly Summary"][f"G{weekly_summary_row}"] = DEFAULT_BASE_HOURS
            replacements["Weekly Summary"][f"H{weekly_summary_row}"] = total_hours
            replacements["Weekly Summary"][f"I{weekly_summary_row}"] = actual["issue_count"]
            replacements["Weekly Summary"][f"J{weekly_summary_row}"] = total_hours - DEFAULT_EXPECTED_HOURS
            weekly_summary_row += 1

    chart_row = 81
    for username in roster:
        implementation_total = sum(
            implementation_for(aggregation, week, username)["hours"]
            for week in weeks
        )
        expected_total = DEFAULT_EXPECTED_HOURS * len(weeks)
        actual_total = implementation_total + DEFAULT_BASE_HOURS * len(weeks)
        replacements["Charts"][f"B{chart_row}"] = expected_total
        replacements["Charts"][f"C{chart_row}"] = actual_total
        replacements["Charts"][f"D{chart_row}"] = actual_total - expected_total
        replacements["Charts"][f"E{chart_row}"] = actual_total / expected_total if expected_total else 0
        chart_row += 1

    weekly_total_row = 81
    for week in weeks:
        implementation_total = sum(
            implementation_for(aggregation, week, username)["hours"]
            for username in roster
        )
        replacements["Charts"][f"H{weekly_total_row}"] = implementation_total + DEFAULT_BASE_HOURS * len(roster)
        weekly_total_row += 1

    source_row = 104
    for username in roster:
        for week in weeks:
            implementation_hours = implementation_for(aggregation, week, username)["hours"]
            replacements["Charts"][f"C{source_row}"] = DEFAULT_EXPECTED_HOURS
            replacements["Charts"][f"D{source_row}"] = implementation_hours + DEFAULT_BASE_HOURS
            source_row += 1

    return replacements


def apply_cached_formula_values(xlsx_path, weeks, roster, aggregation):
    replacements_by_sheet = cached_formula_replacements(weeks, roster, aggregation)
    if not replacements_by_sheet:
        return

    sheet_paths = workbook_sheet_paths(xlsx_path)
    path_replacements = {}
    worksheet_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", worksheet_ns)

    with ZipFile(xlsx_path) as archive:
        for sheet_name, cell_values in replacements_by_sheet.items():
            sheet_path = sheet_paths.get(sheet_name)
            if not sheet_path:
                continue
            root = ET.fromstring(archive.read(sheet_path))
            for cell in root.findall(f".//{{{worksheet_ns}}}c"):
                reference = cell.attrib.get("r")
                if reference not in cell_values:
                    continue
                value = cell.find(f"{{{worksheet_ns}}}v")
                if value is None:
                    value = ET.SubElement(cell, f"{{{worksheet_ns}}}v")
                value.text = cached_number(cell_values[reference])
            path_replacements[sheet_path] = ET.tostring(root, encoding="utf-8", xml_declaration=True)

    temp_path = xlsx_path.with_name(f".{xlsx_path.name}.cached-values.tmp")
    with ZipFile(xlsx_path) as source, ZipFile(temp_path, "w", ZIP_DEFLATED) as target:
        for info in source.infolist():
            target.writestr(info, path_replacements.get(info.filename, source.read(info.filename)))
    shutil.move(temp_path, xlsx_path)


def main():
    if len(sys.argv) not in (2, 3):
        usage()

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    output_path = Path(sys.argv[2]) if len(sys.argv) == 3 else input_path.with_suffix(".xlsx")
    data = json.loads(input_path.read_text())

    weeks = list(data["context"]["report_weeks"])
    roster = roster_from_data(data)
    aggregation = aggregate_rows(data)

    workbook = Workbook()
    add_readme(workbook, data, weeks)
    add_settings(workbook, data)
    add_people(workbook, roster)
    add_base_hours(workbook, weeks, roster)

    for week in weeks:
        add_week_sheet(workbook, week, roster, aggregation)

    add_weekly_summary(workbook, weeks, roster)
    add_chart(workbook, weeks, roster, aggregation)
    add_manual_review(workbook, data)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    style_workbook(workbook)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = "0.00"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    apply_cached_formula_values(output_path, weeks, roster, aggregation)
    print(output_path)


if __name__ == "__main__":
    main()
